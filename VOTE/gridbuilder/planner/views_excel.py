"""
Excel export for the planner grid.
Content of the workbook will be defined by product requirements.
"""
import io
import re
from datetime import date, datetime, time

from django.conf import settings
from django.db import IntegrityError, connection, transaction
from django.db.models import Max
from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintOptions
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties

from .models import CalendarColumnHeader, PlannerExport


def _event_code(request):
    """Current event (EMEA/NA) from session; default EMEA for guests."""
    return request.session.get("event_code", "EMEA")


def _event_year_for_export(request):
    y = request.session.get("event_year")
    if y is not None and str(y).strip():
        return str(y).strip()[:4]
    return str(date.today().year)


def _export_filename_date_part(d):
    """yyyy_mm_dd segment for grid export download filename."""
    return f"{d.year}_{d.month:02d}_{d.day:02d}"


def _export_footer_date_str(d):
    """yyyy/mm/dd for Excel footer (same calendar date as filename)."""
    return d.strftime("%Y/%m/%d")


def _allocate_export_version(event_code, event_year_str):
    """
    Next footer version: max ``event_version`` for this event_code + 1 (NULL/missing → 1).
    Persists that value on all ``planner_export`` rows for ``event_code``, or inserts one row.
    """
    with transaction.atomic():
        v = PlannerExport.objects.filter(event_code=event_code).aggregate(
            m=Max("event_version")
        )["m"]
        version_num = 1 if v is None else v + 1

        updated = PlannerExport.objects.filter(event_code=event_code).update(
            event_version=version_num
        )
        if updated == 0:
            try:
                PlannerExport.objects.create(
                    event_code=str(event_code).strip()[:10],
                    session_event=str(event_code).strip()[:10],
                    event_year=str(event_year_str).strip()[:4],
                    event_version=version_num,
                )
            except IntegrityError:
                PlannerExport.objects.filter(event_code=event_code).update(
                    event_version=version_num
                )

        return version_num


def _apply_grid_sheet_header_footer(ws_grid, header_center, footer_center):
    """Header/footer center text on GridBySpeakerType only."""
    ws_grid.HeaderFooter.scaleWithDoc = False
    ws_grid.HeaderFooter.alignWithMargins = False
    ws_grid.oddHeader.center.text = header_center
    ws_grid.oddFooter.center.text = footer_center


# Print layout aligned with legacy ``IDUG_EMEA2025-Grid_*.xlsx`` → ``GridBySpeakerType`` sheet
# (fitToPage + scale % + fitToHeight 0 + full-width row breaks — not fitToWidth).
EXCEL_GRID_PRINT_PAPER_SIZE = 8
EXCEL_GRID_PRINT_SCALE = 54
EXCEL_GRID_PAGE_MARGINS_INCH = dict(
    left=0.47244094488188981,
    right=0.47244094488188981,
    top=0.94488188976377963,
    bottom=0.55118110236220474,
    header=0.39370078740157483,
    footer=0.22647058823529412,
)


def _apply_grid_sheet_print_layout(
    ws_grid, grid_last_col, date_rows, alternates_header_row, event_code
):
    """
    Grid sheet only: print area B2 through last track column and last used row.
    Manual horizontal page breaks: **before the Alternates title row** for all events; and before
    the **third** grey date row for **NA** logins, **fourth** date row for **EMEA** (and others).

    OOXML ``brk@id`` is **zero-based** (break above that row); openpyxl rows are 1-based →
    ``id = excel_row - 1``. Row breaks use full sheet column span (``max=16383``) like the
    reference workbook—not a narrow min/max band.

    Page setup matches the working legacy grid file: ``fitToPage``, ``scale`` %, ``fitToHeight=0``,
    **no** ``fitToWidth`` (that combination hides breaks). Margins and print options match that file.
    """
    if grid_last_col is None:
        return
    last_row = max(ws_grid.max_row or 2, 2)
    end_letter = get_column_letter(grid_last_col)
    ws_grid.print_area = f"$B$2:${end_letter}${last_row}"

    ws_grid.row_breaks.brk = []
    break_rows = []
    sorted_dates = sorted(date_rows)
    code = (event_code or "").strip().upper()
    if code == "NA":
        if len(sorted_dates) >= 3:
            break_rows.append(sorted_dates[2])
    else:
        if len(sorted_dates) >= 4:
            break_rows.append(sorted_dates[3])
    if alternates_header_row is not None:
        break_rows.append(alternates_header_row)
    for rid in sorted(set(break_rows)):
        if rid < 1:
            continue
        ws_grid.row_breaks.append(Break(id=rid - 1, max=16383, man=True))

    ws_grid.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws_grid.page_setup.paperSize = EXCEL_GRID_PRINT_PAPER_SIZE
    ws_grid.page_setup.scale = EXCEL_GRID_PRINT_SCALE
    ws_grid.page_setup.fitToHeight = 0
    ws_grid.page_setup.fitToWidth = None
    ws_grid.page_setup.errors = "blank"

    ws_grid.print_options = PrintOptions(horizontalCentered=True, gridLines=True)
    m = EXCEL_GRID_PAGE_MARGINS_INCH
    ws_grid.page_margins = PageMargins(
        left=m["left"],
        right=m["right"],
        top=m["top"],
        bottom=m["bottom"],
        header=m["header"],
        footer=m["footer"],
    )


# Column headers start at column G (7), width ~4.70" (≈47 character units)
COLUMN_HEADER_START_COL = 7
COLUMN_HEADER_WIDTH = 47

# Sheet "print": column A = Submissions column AJ (Print); width ~4.88", row height ~1.80"
PRINT_SHEET_COL_A_WIDTH_CHARS = (4.88 / 4.70) * COLUMN_HEADER_WIDTH
PRINT_SHEET_ROW_HEIGHT_PT = 1.80 * 72
ALIGN_PRINT_SHEET_A = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
    indent=4,
)

# Row headers: columns A–E, widths in character units (~0.49", 3.66", 0.62", 0.19", 0.62")
ROW_HEADER_WIDTHS = (5, 37, 6, 2, 6)  # A, B, C, D, E
ROW_HEADER_START_ROW = 5
# Row heights in pt: slot 1.06", break 0.19", date row 0.40"
ROW_HEIGHT_SLOT_PT = 1.06 * 72
ROW_HEIGHT_BREAK_PT = 0.19 * 72
ROW_HEIGHT_DATE_PT = 0.40 * 72

# Row 4: C4="Start", E4="End", then row 4 hidden; column F hidden
ROW_4_HEADER_ROW = 4
COLUMN_F_INDEX = 6

# Tags helpers: fixed block starting at R; one column per track column (G..L → R..W).
GRID_TAGS_HELPER_START_COL = 18

# Calendar extends from column B to last column header (e.g. G..L)
# Styling: date row = gray bg + white bold; break row = light gray bg
FILL_DATE_ROW = PatternFill(patternType="solid", fgColor="808080")
FILL_BREAK_ROW = PatternFill(patternType="solid", fgColor="D9D9D9")
FONT_DATE_ROW = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_ALTERNATES = Font(name="Calibri", size=11, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
# Column B / track columns: Excel `indent` is character-width steps (0–255), NOT points.
# Values 10 and 42 were far too large in the UI (~100+ pt); use small steps (1 and ~4×).
ALIGN_INDENT_B = Alignment(horizontal="left", indent=1, vertical="center")
ALIGN_TRACK_GRID = Alignment(
    horizontal="left", indent=1, vertical="center", wrap_text=True
)
# Partial colspan merges in track area (not full G..last row — those stay ALIGN_TRACK_GRID).
ALIGN_MERGED_TRACK_CELL = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
ALIGN_VERTICAL_CENTER = Alignment(vertical="center")

# Track columns (G..): CF when cell text contains exact tag (order = first match wins via stopIfTrue)
GRID_CF_TAG_FILLS = (
    ("{Rocket Software}", "92D050"),
    ("{BMC}", "FFE699"),
    ("{Broadcom}", "EDBBA7"),
    ("{IBM}", "B4C7E7"),
    ("{User/Consultant}", "C5E0B4"),
    ("{Vendor}", "EDEDED"),
)
# Tags colors: COUNTIF on hidden helpers (R..; same width count as G..L, e.g. R..W for six tracks).
# Order (stopIfTrue): Keynote → Psp → Expert, then GRID_CF_TAG_FILLS.
GRID_CF_SUBMISSIONS_TAGS = (
    ("Keynote", "FF6600"),
    ("Psp", "E0C2CD"),
    ("Expert", "FFFFFF"),
)
THIN_SIDE = Side(style="thin")

# Worksheet default zoom (Excel normal view) when opening the file
EXCEL_SHEET_ZOOM_GRID_AND_SUBMISSIONS = 85
EXCEL_SHEET_ZOOM_PRINT = 100

# Submissions sheet column indices (1-based)
SUBMISSIONS_COL_GRIDCELL = 2  # B
SUBMISSIONS_COL_FIRST_TIME = 24  # X Speaker 1: First Time
SUBMISSIONS_COL_RATING = 31  # AE
SUBMISSIONS_COL_PRINT = 36  # AJ
SUBMISSIONS_COL_CPC_GRIDCELL = 37  # AK


def _cf_solid_fill(rgb_hex):
    """
    Solid fill for conditional formatting. Set fg and bg to the same RGB so Excel shows
    one full background color (not separate pattern vs background in the Format dialog).
    rgb_hex: 6 hex chars e.g. FF6600 (no leading #).
    """
    h = (rgb_hex or "").strip().lstrip("#")
    return PatternFill(patternType="solid", fgColor=h, bgColor=h)


# Grid upgrade: resolve labels like "A1" = column where row 1 has "A" ∩ row where column A has "1"
GRID_HEADER_ROW = 1
GRID_LABEL_COL_A = 1  # column A: optional row index for A1-style keys
GRID_LABEL_COL_B = 2  # column B: slot label (e.g. D3), matches planner_calendarlayout.label
LABEL_COORD_RE = re.compile(r"^([A-Za-z])(\d+)$")


def _normalize_row_key(v):
    """Normalize values in column A for numeric row keys (1 / 1.0 / '1')."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and not float(v).is_integer():
            return str(v).strip()
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def _build_grid_coordinate_maps(ws_grid, num_header_cols):
    """
    From the generated GridBySpeakerType sheet:
    - Row GRID_HEADER_ROW: track letter → column (G..).
    - Column B: slot label (e.g. D3) → row.
    - Column A: row key for A1-style fallback (e.g. '1' → row 12).
    - Column A: full cell text → row (first row wins), for labels like 'K1' when row 1 has no 'K'.
    """
    last_col = COLUMN_HEADER_START_COL + num_header_cols - 1
    letter_to_col = {}
    for c in range(COLUMN_HEADER_START_COL, last_col + 1):
        v = ws_grid.cell(row=GRID_HEADER_ROW, column=c).value
        if v is None or v == "":
            continue
        key = str(v).strip().upper()
        if key:
            letter_to_col[key[0]] = c

    label_to_row = {}
    number_to_row = {}
    col_a_full_to_row = {}
    max_r = ws_grid.max_row or ROW_HEADER_START_ROW
    for r in range(ROW_HEADER_START_ROW, max_r + 1):
        bv = ws_grid.cell(row=r, column=GRID_LABEL_COL_B).value
        if bv is not None and str(bv).strip() != "":
            label_to_row[str(bv).strip()] = r
        av = ws_grid.cell(row=r, column=GRID_LABEL_COL_A).value
        k = _normalize_row_key(av)
        if k:
            number_to_row[k] = r
        if av is not None and str(av).strip() != "":
            full_a = str(av).strip()
            if full_a not in col_a_full_to_row:
                col_a_full_to_row[full_a] = r

    return letter_to_col, label_to_row, number_to_row, col_a_full_to_row, last_col


def _resolve_grid_target_col_row(
    row_dict, letter_to_col, label_to_row, number_to_row, col_a_full_to_row
):
    """
    Prefer track + layout label (column B on sheet).
    Fallback: parse label as 'A1' → letter + number (column A row index + row 1 letter).
    Last resort: entire label text appears in column A → row; column = first track (G).
    (e.g. 'K1' in A10 when row 1 has no 'K' → cell G10; merges unchanged.)
    """
    track = (row_dict.get("track") or "").strip().upper()[:1]
    label = (row_dict.get("label") or "").strip()
    col = letter_to_col.get(track) if track else None
    row_idx = label_to_row.get(label) if label else None
    if col is not None and row_idx is not None:
        return col, row_idx
    m = LABEL_COORD_RE.match(label)
    if m:
        letter, num = m.group(1).upper(), m.group(2)
        col = letter_to_col.get(letter)
        row_idx = number_to_row.get(num)
        if col is not None and row_idx is not None:
            return col, row_idx
    if label and label in col_a_full_to_row:
        return COLUMN_HEADER_START_COL, col_a_full_to_row[label]
    return None, None


def _db_schema():
    """Schema from DATABASES['default'] OPTIONS (e.g. grid) for qualified table names."""
    opts = settings.DATABASES.get("default", {}).get("OPTIONS", {})
    opt_str = opts.get("options", "")
    m = re.search(r"search_path=(\w+)", opt_str)
    return m.group(1) if m else "public"


def _excel_formula_string_literal(s):
    """Double-quoted string for use inside an Excel formula; internal " → ""."""
    return '"' + str(s).replace('"', '""') + '"'


def _grid_cpc_cell_formula(label, submissions_sheet_title, submissions_last_row):
    """CPC sheet: rating line 1, then VLOOKUP slot → Submissions CPCGridCell (AK)."""
    lit = _excel_formula_string_literal(label)
    safe_title = str(submissions_sheet_title).replace("'", "''")
    lr = int(submissions_last_row)
    ref_rating = f"'{safe_title}'!$A$2:$AE${lr}"
    ref_cpc = f"'{safe_title}'!$A$2:$AK${lr}"
    rating = (
        f"IFERROR(TEXT(VLOOKUP({lit},{ref_rating},{SUBMISSIONS_COL_RATING},FALSE),"
        f'"0.00")&CHAR(10),"")'
    )
    body = f"VLOOKUP({lit},{ref_cpc},{SUBMISSIONS_COL_CPC_GRIDCELL},FALSE)"
    return f"={rating}&{body}"


def _apply_cpc_sheet_decorations(
    ws_cpc, event_code, num_header_cols, submissions_sheet_title, submissions_last_row
):
    """
    CPC: rating on line 1 + VLOOKUP to Submissions CPCGridCell (column AK).
    Sheet 1 is unchanged; only ws_cpc is modified.
    """
    if submissions_last_row is None or submissions_last_row < 2:
        return
    letter_to_col, label_to_row, number_to_row, col_a_full_to_row, _last_col = (
        _build_grid_coordinate_maps(ws_cpc, num_header_cols)
    )
    rows = _fetch_layout_grid_upgrade_rows(event_code)
    for row in rows:
        if not row.get("visible", True):
            continue
        if (row.get("type") or "").strip().lower() == "break":
            continue
        col, r = _resolve_grid_target_col_row(
            row, letter_to_col, label_to_row, number_to_row, col_a_full_to_row
        )
        if col is None or r is None:
            continue
        label = row.get("label")
        if label in (None, ""):
            continue
        key = str(label).strip()
        cell = ws_cpc.cell(row=r, column=col)
        val = cell.value
        if val is not None and str(val).lstrip().startswith("="):
            cell.value = _grid_cpc_cell_formula(
                key, submissions_sheet_title, submissions_last_row
            )


def _grid_submissions_lookup_formula(label, submissions_sheet_title, last_row):
    """
    IFNA(VLOOKUP(label, Submissions!A:B col 2), label) — GridCell from Submissions or label if #N/A.
    last_row: end row for $A$2:$B$last_row (at least 2).
    """
    lit = _excel_formula_string_literal(label)
    safe_title = str(submissions_sheet_title).replace("'", "''")
    ref = f"'{safe_title}'!$A$2:$B${int(last_row)}"
    return f"=VLOOKUP({lit},{ref},2,FALSE)"


def _grid_submissions_tags_helper_formula(label, submissions_sheet_title, last_row):
    """Nested VLOOKUP(label→GridCell, then GridCell→Tags col 5). Same label literal as track cell."""
    lit = _excel_formula_string_literal(label)
    safe_title = str(submissions_sheet_title).replace("'", "''")
    lr = int(last_row)
    ref_ab = f"'{safe_title}'!$A$2:$B${lr}"
    ref_bf = f"'{safe_title}'!$B$2:$F${lr}"
    return f'=VLOOKUP(VLOOKUP({lit},{ref_ab},2,FALSE),{ref_bf},5,FALSE)'


def _fetch_layout_grid_upgrade_rows(event_code):
    """
    Same shape as product query; `track` included so we can resolve column + row (track + label).
    event_code from login (EMEA / NA).
    """
    schema = _db_schema()
    sql = f"""
    SELECT l.day,
           l.colspan,
           l.visible,
           l.type,
           l.label,
           l.track
    FROM "{schema}".planner_calendarlayout l
    WHERE l.event_code = %s
      AND l.type = 'slot'
      AND l.label IS NOT NULL
    ORDER BY l.day, l.time_slot_id, l.track
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [event_code])
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _upgrade_grid_from_layout_sql(
    ws_grid, event_code, num_header_cols, submissions_sheet_title, submissions_last_row
):
    """
    After the grid sheet is built: resolve each layout row to a calendar cell (track + label,
    or 'A1' style, or full label in column A → column G), write IFNA(VLOOKUP→Submissions GridCell),
    merge colspan. Skips visible=false. Skips break rows (already merged in phase 1 grid build).
    """
    letter_to_col, label_to_row, number_to_row, col_a_full_to_row, last_col = (
        _build_grid_coordinate_maps(ws_grid, num_header_cols)
    )
    rows = _fetch_layout_grid_upgrade_rows(event_code)
    for row in rows:
        if not row.get("visible", True):
            continue
        row_type = (row.get("type") or "").strip().lower()
        if row_type == "break":
            continue
        col, r = _resolve_grid_target_col_row(
            row, letter_to_col, label_to_row, number_to_row, col_a_full_to_row
        )
        if col is None or r is None:
            continue
        colspan = row.get("colspan")
        try:
            cs = int(colspan) if colspan is not None else 1
        except (TypeError, ValueError):
            cs = 1
        if cs < 1:
            cs = 1
        end_col = min(col + cs - 1, last_col)
        val = row.get("label") if row.get("label") not in (None, "") else row.get("type")
        if cs > 1:
            try:
                ws_grid.merge_cells(
                    start_row=r,
                    start_column=col,
                    end_row=r,
                    end_column=end_col,
                )
            except ValueError:
                pass
        top_left = ws_grid.cell(row=r, column=col)
        top_left.value = _grid_submissions_lookup_formula(
            val, submissions_sheet_title, submissions_last_row
        )

        if submissions_last_row is not None and submissions_last_row >= 2:
            span = end_col - col
            # G→R, H→S, … (same index within the G..L and R..W blocks).
            h_start = GRID_TAGS_HELPER_START_COL + (col - COLUMN_HEADER_START_COL)
            h_end = h_start + span
            formula = _grid_submissions_tags_helper_formula(
                val, submissions_sheet_title, submissions_last_row
            )
            # One formula per helper column in the span (not merged): each S,T,… holds text so
            # track column H,I,… CF can read its parallel helper; merges would leave S,T empty.
            for hc in range(h_start, h_end + 1):
                ws_grid.cell(row=r, column=hc).value = formula


def _hide_grid_tags_helper_columns(ws_grid, num_header_cols):
    """Hide GRID_TAGS_HELPER_START_COL … for num_header_cols (e.g. R..W when track is G..L)."""
    for i in range(num_header_cols):
        letter = get_column_letter(GRID_TAGS_HELPER_START_COL + i)
        ws_grid.column_dimensions[letter].hidden = True


def _row_a_has_value(v):
    """True if Excel column A should be treated as having a value for alignment rules."""
    if v is None:
        return False
    if isinstance(v, bool):
        return True
    if isinstance(v, (int, float)):
        return True
    return str(v).strip() != ""


def _merge_anchor_is_formula(ws_grid, mr):
    """True if merged region top-left holds an Excel formula (layout/keynote); breaks use plain text."""
    if mr is None:
        return False
    val = ws_grid.cell(row=mr.min_row, column=mr.min_col).value
    if val is None:
        return False
    s = str(val).lstrip()
    return s.startswith("=")


def _apply_grid_alignment(ws_grid, num_header_cols):
    """
    Column B: left align (all rows).
    Columns G..last where column A has a value: unmerged → ALIGN_TRACK_GRID; merged colspan
    (not spanning full G..last on that row) → centered H+V. Full-width track merges (G..last):
    breaks stay left (plain text); keynotes / layout slots stay centered (formula in anchor cell).
    Run after grid body + layout upgrade so merged cells get alignment.
    """
    last_col = COLUMN_HEADER_START_COL + num_header_cols - 1
    max_r = ws_grid.max_row or 1
    for r in range(1, max_r + 1):
        ws_grid.cell(row=r, column=2).alignment = ALIGN_INDENT_B
        if not _row_a_has_value(ws_grid.cell(row=r, column=1).value):
            continue
        for c in range(COLUMN_HEADER_START_COL, last_col + 1):
            mr = _merged_range_containing(ws_grid, r, c)
            if mr is None:
                ws_grid.cell(row=r, column=c).alignment = ALIGN_TRACK_GRID
                continue
            if (r, c) != (mr.min_row, mr.min_col):
                continue
            spans_multiple = (mr.max_col > mr.min_col) or (mr.max_row > mr.min_row)
            if not spans_multiple:
                ws_grid.cell(row=r, column=c).alignment = ALIGN_TRACK_GRID
                continue
            full_track_row = (
                mr.min_row == mr.max_row == r
                and mr.min_col == COLUMN_HEADER_START_COL
                and mr.max_col == last_col
            )
            if full_track_row:
                if _merge_anchor_is_formula(ws_grid, mr):
                    ws_grid.cell(row=r, column=c).alignment = ALIGN_MERGED_TRACK_CELL
                else:
                    ws_grid.cell(row=r, column=c).alignment = ALIGN_TRACK_GRID
            else:
                ws_grid.cell(row=r, column=c).alignment = ALIGN_MERGED_TRACK_CELL


def _apply_grid_track_submissions_tags_conditional_formatting(
    ws_grid, num_header_cols, submissions_sheet_title, submissions_last_row
):
    """
    Excel: COUNTIF on the hidden helper aligned with each track column (G→R, H→S, …).
    Uses ADDRESS(ROW(),COLUMN()+offset) so each cell in G..L reads its parallel R..W cell.
    """
    if submissions_last_row is None or submissions_last_row < 2:
        return
    last_col = COLUMN_HEADER_START_COL + num_header_cols - 1
    first_row = ROW_HEADER_START_ROW
    last_row = ws_grid.max_row or first_row
    if last_row < first_row:
        return

    start_letter = get_column_letter(COLUMN_HEADER_START_COL)
    end_letter = get_column_letter(last_col)
    cell_range = f"{start_letter}{first_row}:{end_letter}{last_row}"
    col_off = GRID_TAGS_HELPER_START_COL - COLUMN_HEADER_START_COL

    for search_term, fg in GRID_CF_SUBMISSIONS_TAGS:
        st = search_term.replace('"', '""')
        formula = (
            f'=COUNTIF(INDIRECT(ADDRESS(ROW(),COLUMN()+{col_off})),"*{st}*")>0'
        )
        fill = _cf_solid_fill(fg)
        ws_grid.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[formula], fill=fill, stopIfTrue=True),
        )


def _apply_grid_track_conditional_formatting(ws_grid, num_header_cols):
    """
    Excel: track area G..last column, from first calendar row to last used row.
    Text contains tag (COUNTIF wildcards); first matching rule fills the cell (stopIfTrue).
    Apply after _apply_grid_track_submissions_tags_conditional_formatting.
    """
    last_col = COLUMN_HEADER_START_COL + num_header_cols - 1
    first_row = ROW_HEADER_START_ROW
    last_row = ws_grid.max_row or first_row
    if last_row < first_row:
        return

    start_letter = get_column_letter(COLUMN_HEADER_START_COL)
    end_letter = get_column_letter(last_col)
    cell_range = f"{start_letter}{first_row}:{end_letter}{last_row}"
    first_cell = f"{start_letter}{first_row}"

    for token, fg in GRID_CF_TAG_FILLS:
        safe = token.replace('"', '""')
        formula = f'=COUNTIF({first_cell},"*{safe}*")>0'
        fill = _cf_solid_fill(fg)
        ws_grid.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[formula], fill=fill, stopIfTrue=True),
        )


def _write_grid_column_headers(ws_grid, event_code):
    """Write column headers (track, subject, room_name) from planner_calendarcolumnheader.
    Columns start at G; row 1=track (Calibri 11), row 2=subject, row 3=room_name (Calibri 16 bold center).
    Returns number of header columns (calendar last column = COLUMN_HEADER_START_COL + this - 1).
    """
    headers = list(
        CalendarColumnHeader.objects.filter(event_code=event_code)
        .order_by("id")
        .values("track", "subject", "room_name")
    )
    font_track = Font(name="Calibri", size=11)
    font_subject_room = Font(name="Calibri", size=16, bold=True)
    align_center = Alignment(horizontal="center")

    # Row 2 and 3 height: 0.26" ≈ 18.72 pt
    ws_grid.row_dimensions[2].height = 0.26 * 72
    ws_grid.row_dimensions[3].height = 0.26 * 72

    for i, h in enumerate(headers):
        col = COLUMN_HEADER_START_COL + i
        letter = get_column_letter(col)
        ws_grid.column_dimensions[letter].width = COLUMN_HEADER_WIDTH

        ws_grid.cell(row=1, column=col, value=h["track"]).font = font_track
        cell_s = ws_grid.cell(row=2, column=col, value=h["subject"])
        cell_s.font = font_subject_room
        cell_s.alignment = align_center
        cell_r = ws_grid.cell(row=3, column=col, value=h["room_name"])
        cell_r.font = font_subject_room
        cell_r.alignment = align_center

    return len(headers)


def _fetch_row_header_rows(event_code):
    """Run the row-headers query; returns one row per time slot (day, order, type, track, day_label, label, start_time, end_time)."""
    sql = """
    SELECT DISTINCT s.day, s.order, l.type,
           CASE WHEN LEFT(l.label, 1) = 'K' THEN l.label ELSE SUBSTRING(l.label FROM 2) END AS track,
           d.day AS day_label, s.label, s.start_time, s.end_time
    FROM planner_calendartimeslot s
    INNER JOIN planner_day d ON d.id = s.day
    LEFT JOIN planner_calendarlayout l ON l.time_slot_id = s.id
    WHERE s.event_code = %s AND l.visible = true AND l.type IN ('slot', 'break')
    ORDER BY s.day, s.order
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, [event_code])
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _apply_fill_range(ws_grid, row, col_start, col_end, fill):
    """Apply fill from col_start to col_end (inclusive) on the given row."""
    for c in range(col_start, col_end + 1):
        ws_grid.cell(row=row, column=c).fill = fill


# Row header columns for per-row outline blocks (B–E)
ROW_HEADER_COL_B = 2
ROW_HEADER_COL_E = 5


def _merged_range_containing(ws, row, col):
    """Return the CellRange that contains (row, col), or None."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None


def _add_border_sides(ws, row, col, top=None, bottom=None, left=None, right=None):
    """Merge new thin sides onto existing cell border."""
    cell = ws.cell(row=row, column=col)
    prev = cell.border
    cell.border = Border(
        top=top or prev.top,
        bottom=bottom or prev.bottom,
        left=left or prev.left,
        right=right or prev.right,
    )


def _apply_frame_edge_top(ws, r1, c1, c2):
    """Top perimeter at row r1; merge-aware (top on merge top-left when merge is one row at r1)."""
    c = c1
    while c <= c2:
        mr = _merged_range_containing(ws, r1, c)
        if mr is None:
            _add_border_sides(ws, r1, c, top=THIN_SIDE)
            c += 1
            continue
        if (r1, c) != (mr.min_row, mr.min_col):
            c += 1
            continue
        if mr.min_row != r1:
            c += 1
            continue
        if mr.max_row != r1:
            c += 1
            continue
        _add_border_sides(ws, mr.min_row, mr.min_col, top=THIN_SIDE)
        c = mr.max_col + 1


def _apply_frame_edge_bottom(ws, r2, c1, c2):
    """Bottom perimeter at row r2; merge-aware."""
    c = c1
    while c <= c2:
        mr = _merged_range_containing(ws, r2, c)
        if mr is None:
            _add_border_sides(ws, r2, c, bottom=THIN_SIDE)
            c += 1
            continue
        if (r2, c) != (mr.min_row, mr.min_col):
            c += 1
            continue
        if mr.min_row != r2:
            c += 1
            continue
        if mr.max_row != r2:
            c += 1
            continue
        _add_border_sides(ws, mr.min_row, mr.min_col, bottom=THIN_SIDE)
        c = mr.max_col + 1


def _apply_frame_edge_left(ws, r1, r2, c1):
    """Left perimeter at col c1; merge-aware."""
    r = r1
    while r <= r2:
        mr = _merged_range_containing(ws, r, c1)
        if mr is None:
            _add_border_sides(ws, r, c1, left=THIN_SIDE)
            r += 1
            continue
        if mr.min_col != c1:
            r += 1
            continue
        if (r, c1) != (mr.min_row, mr.min_col):
            r += 1
            continue
        _add_border_sides(ws, mr.min_row, mr.min_col, left=THIN_SIDE)
        r = mr.max_row + 1


def _apply_frame_edge_right(ws, r1, r2, c2):
    """Right perimeter at col c2; merge-aware (right edge on merge top-left when max_col == c2)."""
    r = r1
    while r <= r2:
        mr = _merged_range_containing(ws, r, c2)
        if mr is None:
            _add_border_sides(ws, r, c2, right=THIN_SIDE)
            r += 1
            continue
        if mr.max_col != c2:
            r += 1
            continue
        # (r,c2) may be non-top-left of a horizontal merge ending at c2; Excel shows right border on anchor.
        _add_border_sides(ws, mr.min_row, mr.min_col, right=THIN_SIDE)
        r = mr.max_row + 1


def _apply_outer_frame_four_edges(ws, r1, c1, r2, c2):
    """Draw only the four edges of [r1,c1]–[r2,c2] (Option B), merge-aware for colspan/break rows."""
    _apply_frame_edge_top(ws, r1, c1, c2)
    _apply_frame_edge_bottom(ws, r2, c1, c2)
    _apply_frame_edge_left(ws, r1, r2, c1)
    _apply_frame_edge_right(ws, r1, r2, c2)


def _apply_calendar_rows_extra_thin_outer_border(ws, last_calendar_row, last_col, date_rows=None):
    """
    Final outside border on B..last_col for each main calendar row (e.g. B5:L5, B6:L6, …).
    Date rows are skipped; they get borders from _apply_date_row_outline only.
    Runs after all other border formatting.
    """
    if last_calendar_row is None or last_col is None:
        return
    if last_calendar_row < ROW_HEADER_START_ROW:
        return
    for row in range(ROW_HEADER_START_ROW, last_calendar_row + 1):
        if date_rows is not None and row in date_rows:
            continue
        cell = ws.cell(row=row, column=last_col)
        cell.border = Border(
            top=THIN_SIDE,
            bottom=THIN_SIDE,
            left=THIN_SIDE,
            right=THIN_SIDE,
        )

        #_set_full_outer_rect_border(ws, row, ROW_HEADER_COL_B, row, last_col)


def _apply_date_row_outline(ws, date_rows, last_col):
    """Thin outer rectangle on grey date header rows only (B..last_col); does not alter slot/break rows."""
    if not date_rows or last_col is None:
        return
    for row in date_rows:
        _set_outer_rect_border(ws, row, ROW_HEADER_COL_B, row, last_col)
        # _set_outer_rect_border skips non-anchor cells in merges; outer right at last_col must sit on merge anchor.
        mr = _merged_range_containing(ws, row, last_col)
        if mr is not None and mr.max_col == last_col:
            _add_border_sides(ws, mr.min_row, mr.min_col, right=THIN_SIDE)
        else:
            _add_border_sides(ws, row, last_col, right=THIN_SIDE)


def _set_outer_rect_border(ws, r1, c1, r2, c2, edge_side=THIN_SIDE):
    """Outside border on rectangle [r1,c1]–[r2,c2]. Respects merged cells (outline on merge)."""
    mr = _merged_range_containing(ws, r1, c1)
    if (
        mr is not None
        and mr.min_row == r1
        and mr.min_col == c1
        and mr.max_row == r2
        and mr.max_col == c2
    ):
        cell = ws.cell(row=r1, column=c1)
        cell.border = Border(
            top=edge_side,
            bottom=edge_side,
            left=edge_side,
            right=edge_side,
        )
        return
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            inner = _merged_range_containing(ws, r, c)
            if inner is not None and (r, c) != (inner.min_row, inner.min_col):
                continue
            top = edge_side if r == r1 else None
            btm = edge_side if r == r2 else None
            left = edge_side if c == c1 else None
            right = edge_side if c == c2 else None
            cell = ws.cell(row=r, column=c)
            prev = cell.border
            cell.border = Border(
                top=top or prev.top,
                bottom=btm or prev.bottom,
                left=left or prev.left,
                right=right or prev.right,
            )


def _apply_row_border_track_columns(ws, row, last_col, date_rows):
    """Outside borders for G..last_col on one row; merge-aware. Skips date rows."""
    if row in date_rows:
        return
    c = COLUMN_HEADER_START_COL
    while c <= last_col:
        mr = _merged_range_containing(ws, row, c)
        if mr is None:
            _set_outer_rect_border(ws, row, c, row, c)
            c += 1
            continue
        if mr.min_row != row or mr.max_row != row:
            c = mr.max_col + 1
            continue
        if (row, c) != (mr.min_row, mr.min_col):
            c = mr.max_col + 1
            continue
        _set_outer_rect_border(ws, mr.min_row, mr.min_col, mr.max_row, mr.max_col)
        c = mr.max_col + 1


def _apply_grid_calendar_block_borders(ws, last_calendar_row, last_calendar_col, date_rows):
    """
    B–E: one outside border block per slot/break row. G–last: outside borders (merged colspan).
    Rows: ROW_HEADER_START_ROW..last_calendar_row. Date rows unchanged (no extra borders).
    """
    if last_calendar_row is None or last_calendar_col is None:
        return
    for row in range(ROW_HEADER_START_ROW, last_calendar_row + 1):
        if row in date_rows:
            continue
        _set_outer_rect_border(ws, row, ROW_HEADER_COL_B, row, ROW_HEADER_COL_E)
        _apply_row_border_track_columns(ws, row, last_calendar_col, date_rows)


def _apply_grid_track_header_outer_border(ws, last_col):
    """Outside thin border around track headers: rows 2–3, columns G..last (e.g. G2:L3)."""
    if last_col is None or last_col < COLUMN_HEADER_START_COL:
        return
    _set_outer_rect_border(ws, 2, COLUMN_HEADER_START_COL, 3, last_col)


def _apply_grid_track_header_inner_verticals(ws, last_col):
    """Thin vertical lines between adjacent track header columns on rows 2–3 (inside G..last)."""
    if last_col is None or last_col <= COLUMN_HEADER_START_COL:
        return
    for c in range(COLUMN_HEADER_START_COL, last_col):
        _add_border_sides(ws, 2, c, right=THIN_SIDE)
        _add_border_sides(ws, 3, c, right=THIN_SIDE)


def _apply_grid_calendar_outer_frame(ws, last_calendar_row, last_calendar_col):
    """
    Outside thin border around the main calendar block: B..last_col × ROW_HEADER_START_ROW..last_calendar_row
    (e.g. B5:L58). Option B: four edges only, merge-aware so merged G..L rows show a full perimeter.
    """
    if last_calendar_row is None or last_calendar_col is None:
        return
    if last_calendar_row < ROW_HEADER_START_ROW:
        return
    _apply_outer_frame_four_edges(
        ws,
        ROW_HEADER_START_ROW,
        ROW_HEADER_COL_B,
        last_calendar_row,
        last_calendar_col,
    )


def _write_grid_row_headers(ws_grid, event_code, num_column_headers):
    """Write row headers from row-headers query. Day labels in column B; then per-slot rows A–E.
    last_calendar_col: last column index (1-based) of the calendar (e.g. 12 for L).
    Returns (last_calendar_row, last_calendar_col, date_rows, alternates_header_row).
    date_rows = gray date label rows (no extra borders).
    alternates_header_row = 1-based row of merged "Alternates" title (for print break), or None.
    """
    rows = _fetch_row_header_rows(event_code)
    if not rows:
        return None, None, set(), None

    last_calendar_col = COLUMN_HEADER_START_COL + num_column_headers - 1

    # One row per time slot: keep first occurrence per (day, order)
    seen = set()
    slots = []
    for r in rows:
        key = (r["day"], r["order"])
        if key not in seen:
            seen.add(key)
            slots.append(r)

    font_label_bold = Font(name="Calibri", size=11, bold=True)
    for col_idx, width in enumerate(ROW_HEADER_WIDTHS, start=1):
        ws_grid.column_dimensions[get_column_letter(col_idx)].width = width

    current_row = ROW_HEADER_START_ROW
    prev_day = None
    last_calendar_row = None  # last row inside calendar (before 2 empty lines before Alternates)
    date_rows = set()
    alternates_header_row = None

    for r in slots:
        if r["day"] != prev_day:
            day_label = (r["day_label"] or "").strip()
            if day_label == "Alternates":
                # 2 empty rows, then merged C:D:E:F with "Alternates" only (no date row in B – exception)
                current_row += 2
                alternates_header_row = current_row
                ws_grid.merge_cells(
                    start_row=current_row,
                    start_column=3,
                    end_row=current_row,
                    end_column=6,
                )
                cell_alt = ws_grid.cell(row=current_row, column=3, value="Alternates")
                cell_alt.font = FONT_ALTERNATES
                cell_alt.alignment = ALIGN_CENTER
                current_row += 1
                prev_day = r["day"]
                # fall through to write the first Alternate slot row (and all following)
            else:
                # Normal date row: B = day label, height 0.40", Calibri 14 bold white, bg #808080 to last col
                cell_date = ws_grid.cell(row=current_row, column=2, value=day_label)
                cell_date.font = FONT_DATE_ROW
                cell_date.alignment = ALIGN_INDENT_B
                ws_grid.row_dimensions[current_row].height = ROW_HEIGHT_DATE_PT
                _apply_fill_range(ws_grid, current_row, 2, last_calendar_col, FILL_DATE_ROW)
                last_calendar_row = current_row
                date_rows.add(current_row)
                current_row += 1
                prev_day = r["day"]

        # A=track (empty when null), B=label (bold, indent), C=start_time, D="-" only if both times, E=end_time
        track_val = r["track"] if r["track"] is not None else ""
        start_t = r["start_time"]
        end_t = r["end_time"]
        start_str = start_t.strftime("%H:%M") if start_t else ""
        end_str = end_t.strftime("%H:%M") if end_t else ""
        col_d_val = "-" if (start_t and end_t) else ""

        ws_grid.cell(row=current_row, column=1, value=track_val).alignment = ALIGN_VERTICAL_CENTER
        cell_b = ws_grid.cell(row=current_row, column=2, value=r["label"] or "")
        # Alternates: column B not bold; normal days: bold
        cell_b.font = font_label_bold if (r["day_label"] or "").strip() != "Alternates" else Font(name="Calibri", size=11)
        cell_b.alignment = ALIGN_INDENT_B
        ws_grid.cell(row=current_row, column=3, value=start_str).alignment = ALIGN_VERTICAL_CENTER
        ws_grid.cell(row=current_row, column=4, value=col_d_val).alignment = ALIGN_VERTICAL_CENTER
        ws_grid.cell(row=current_row, column=5, value=end_str).alignment = ALIGN_VERTICAL_CENTER

        row_height = ROW_HEIGHT_SLOT_PT if r["type"] == "slot" else ROW_HEIGHT_BREAK_PT
        ws_grid.row_dimensions[current_row].height = row_height
        if r["type"] == "break":
            _apply_fill_range(ws_grid, current_row, 2, last_calendar_col, FILL_BREAK_ROW)
            # Merge G to last column, value = type (e.g. "break"), centered
            if last_calendar_col >= COLUMN_HEADER_START_COL:
                ws_grid.merge_cells(
                    start_row=current_row,
                    start_column=COLUMN_HEADER_START_COL,
                    end_row=current_row,
                    end_column=last_calendar_col,
                )
                cell_break = ws_grid.cell(row=current_row, column=COLUMN_HEADER_START_COL, value=r["label"] or r["type"])
                cell_break.fill = FILL_BREAK_ROW
                cell_break.alignment = ALIGN_CENTER
        # Border stops before 2 empty rows + Alternates block; include Alternates slot rows in sheet but not in border
        if (r["day_label"] or "").strip() != "Alternates":
            last_calendar_row = current_row
        current_row += 1

    # Vertical center for all cells A5 to last column (borders applied later after layout merges)
    if last_calendar_row is not None:
        for row in range(ROW_HEADER_START_ROW, last_calendar_row + 1):
            for col in range(1, last_calendar_col + 1):
                cell = ws_grid.cell(row=row, column=col)
                if cell.alignment is None or cell.alignment.vertical is None:
                    cell.alignment = ALIGN_VERTICAL_CENTER

    return last_calendar_row, last_calendar_col, date_rows, alternates_header_row


# --- Submissions{YY} sheet (session export for Sessionboard-style columns) ---
SUBMISSIONS_HEADERS = [
    "Slot",
    "GridCell",
    "Friendly ID",
    "Title",
    "Description",
    "Tags",
    "Status",
    "Day",
    "Starts At",
    "Ends At",
    "Location",
    "Submitter",
    "Audience",
    "Learning Objective 1",
    "Session Type",
    "Subject",
    "This presentation ...",
    "Speakers",
    "Speaker 1: First Name",
    "Speaker 1: Last Name",
    "Speaker 1: Job Title",
    "Speaker 1: Company Name",
    "Speaker 1: Email",
    "Speaker 1: First Time",
    "Speaker 2: First Name",
    "Speaker 2: Last Name",
    "Speaker 2: Job Title",
    "Speaker 2: Company Name",
    "Speaker 2: Email",
    "Speaker 2: First Time",
    # Rating header filled dynamically with event + year
    None,  # placeholder index 30 -> column AE
    "Internal Comments",
    "External Comments",
    "Session Type",
    "Event",
    "Print",
    "CPCGridCell",
]


def _submissions_sql_select(event_code):
    """Same shape as product SQL; uses unqualified names (search_path)."""
    return """
    SELECT l.label AS slot,
           s.session_code AS friendly_id,
           s.title,
           s.description,
           string_agg(pc.code, ', ' ORDER BY pc.code) AS tags,
           d.day AS day,
           t.start_time AS start_time,
           t.end_time AS end_time,
           s.submitter,
           s.audience,
           s.objective_1,
           s.session_type AS session_type_session,
           sb.subject_code AS subject,
           s.presentation AS presentation,
           s.speakers AS speakers,
           s.speaker_1_first_name,
           s.speaker_1_last_name,
           s.speaker_1_title,
           s.speaker_1_company,
           s.speaker_1_email,
           s.speaker_1_first_time,
           s.speaker_2_first_name,
           s.speaker_2_last_name,
           s.speaker_2_title,
           s.speaker_2_company,
           s.speaker_2_email,
           s.speaker_2_first_time,
           s.rating AS rating,
           s.internal_comments,
           s.external_comments,
           st.description AS sessiontype,
           s.event_code AS event_code
    FROM planner_session s
    LEFT JOIN planner_calendarslot c ON s.id = c.session_id
    LEFT JOIN planner_calendarlayout l ON l.id = c.layout_id
    LEFT JOIN planner_day d ON d.id = l.day
    LEFT JOIN planner_calendartimeslot t ON t.id = l.time_slot_id
    LEFT JOIN planner_sessiontype st ON st.id = s.session_type_id
    LEFT JOIN planner_subject sb ON sb.subject_id = s.subject_id
    LEFT JOIN planner_session_topic tp ON tp.session_id = s.id
    LEFT JOIN planner_topic pc ON pc.id = tp.topic_id
    WHERE s.event_code = %s
    GROUP BY
      l.label, d.day, t.start_time, t.end_time,
      s.session_code, s.title, s.description, s.submitter,
      s.audience, s.objective_1, s.session_type, sb.subject_code, s.presentation, s.speakers,
      s.speaker_1_first_name, s.speaker_1_last_name, s.speaker_1_title, s.speaker_1_company,
      s.speaker_1_email, s.speaker_1_first_time,
      s.speaker_2_first_name, s.speaker_2_last_name, s.speaker_2_title, s.speaker_2_company,
      s.speaker_2_email, s.speaker_2_first_time,
      s.rating, s.internal_comments, s.external_comments,
      st.description, s.event_code
    ORDER BY s.session_code
    """


def _fmt_time_cell(val):
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    return val


def _fetch_submission_rows(event_code):
    sql = _submissions_sql_select(event_code)
    with connection.cursor() as cursor:
        cursor.execute(sql, [event_code])
        columns = [c[0] for c in cursor.description]
        return columns, list(cursor.fetchall())


def _submissions_formula_suffix(r):
    """Excel fragment: {"&Vr&"}{"&AHr&"}{ }{"&Pr&"}" """
    return '"{"&V%d&"}{"&AH%d&"}{}{"&P%d&"}"' % (r, r, r)


def _submissions_gridcell_formula(r):
    """Column B: GridCell."""
    return (
        f'=A{r}&"("&C{r}&") - "&D{r}&CHAR(10)&" "&S{r}&" "&T{r}&" - "&'
        + _submissions_formula_suffix(r)
    )


def _submissions_cpc_gridcell_formula(r):
    """Column AK: CPCGridCell — GridCell with ⭐ before speaker 1 when column X is Yes."""
    return (
        f'=A{r}&"("&C{r}&") - "&D{r}&CHAR(10)&IF(X{r}="Yes","⭐","")&" "&S{r}&" "&T{r}&" - "&'
        + _submissions_formula_suffix(r)
    )


def _submissions_print_formula(r):
    """Column AJ: Print — rating line, then session block with optional ⭐ (same row)."""
    return (
        f'=IFERROR(TEXT(AE{r},"0.00"),"")&CHAR(10)&"("&C{r}&") - "&D{r}&CHAR(10)&" "&'
        f'IF(X{r}="Yes","⭐","")&S{r}&" "&T{r}&" - "&'
        + _submissions_formula_suffix(r)
    )


def _submissions_special_gridcell_formula(r):
    """Column B for special-session rows: Title if set, else Tags."""
    return f'=IF(D{r}<>"",D{r},F{r})'


def _special_submissions_sql():
    """Calendar slots with a special session type; not all Submissions columns apply."""
    schema = _db_schema()
    return f"""
    SELECT l.label AS slot,
           d.day AS day,
           t.start_time AS start_time,
           t.end_time AS end_time,
           c.description AS title,
           st.name AS tags
    FROM "{schema}".planner_calendarslot c
    LEFT JOIN "{schema}".planner_calendarlayout l ON l.id = c.layout_id
    LEFT JOIN "{schema}".planner_day d ON d.id = l.day
    LEFT JOIN "{schema}".planner_calendartimeslot t ON t.id = l.time_slot_id
    LEFT JOIN "{schema}".planner_specialsessiontype st ON st.id = c.special_session_type_id
    WHERE c.special_session_type_id IS NOT NULL
      AND c.event_code = %s
    ORDER BY d.day, t.start_time, l.label
    """


def _fetch_special_submission_rows(event_code):
    sql = _special_submissions_sql()
    with connection.cursor() as cursor:
        cursor.execute(sql, [event_code])
        return list(cursor.fetchall())


def _write_submissions_sheet(ws_sub, event_code):
    """Fill Submissions sheet: headers, data, formulas B/AK/AJ. Returns last row for $A$2:$B$l (at least 2).
    """
    year = date.today().year
    rating_header = f"Rating: {event_code} {year}"

    headers = list(SUBMISSIONS_HEADERS)
    headers[30] = rating_header
    for col_idx, text in enumerate(headers, start=1):
        ws_sub.cell(row=1, column=col_idx, value=text)

    ws_sub.freeze_panes = "A2"
    ws_sub.auto_filter.ref = "A1:AK1"

    _, rows = _fetch_submission_rows(event_code)
    # Column map from SQL row tuple order (matches SELECT list)
    # slot, friendly_id, title, description, tags, day, start_time, end_time, submitter, audience,
    # objective_1, session_type_session, subject, presentation, speakers, sp1..6, sp2..6, rating, int, ext, sessiontype, event_code

    for i, row_tuple in enumerate(rows, start=0):
        r = i + 2  # Excel row
        (
            slot,
            friendly_id,
            title,
            description,
            tags,
            day,
            start_time,
            end_time,
            submitter,
            audience,
            objective_1,
            session_type_sess,
            subject,
            presentation,
            speakers,
            s1_fn,
            s1_ln,
            s1_title,
            s1_co,
            s1_em,
            s1_ft,
            s2_fn,
            s2_ln,
            s2_title,
            s2_co,
            s2_em,
            s2_ft,
            rating,
            int_comm,
            ext_comm,
            sessiontype,
            ev_code,
        ) = row_tuple

        ws_sub.cell(row=r, column=1, value=slot)
        ws_sub.cell(row=r, column=2, value=_submissions_gridcell_formula(r))
        ws_sub.cell(row=r, column=3, value=friendly_id)
        ws_sub.cell(row=r, column=4, value=title)
        ws_sub.cell(row=r, column=5, value=description)
        ws_sub.cell(row=r, column=6, value=tags)
        ws_sub.cell(row=r, column=7, value=None)
        ws_sub.cell(row=r, column=8, value=day)
        ws_sub.cell(row=r, column=9, value=_fmt_time_cell(start_time))
        ws_sub.cell(row=r, column=10, value=_fmt_time_cell(end_time))
        ws_sub.cell(row=r, column=11, value=None)
        ws_sub.cell(row=r, column=12, value=submitter)
        ws_sub.cell(row=r, column=13, value=audience)
        ws_sub.cell(row=r, column=14, value=objective_1)
        ws_sub.cell(row=r, column=15, value=session_type_sess)
        ws_sub.cell(row=r, column=16, value=subject)
        ws_sub.cell(row=r, column=17, value=presentation)
        ws_sub.cell(row=r, column=18, value=speakers)
        ws_sub.cell(row=r, column=19, value=s1_fn)
        ws_sub.cell(row=r, column=20, value=s1_ln)
        ws_sub.cell(row=r, column=21, value=s1_title)
        ws_sub.cell(row=r, column=22, value=s1_co)
        ws_sub.cell(row=r, column=23, value=s1_em)
        ws_sub.cell(row=r, column=24, value=s1_ft)
        ws_sub.cell(row=r, column=25, value=s2_fn)
        ws_sub.cell(row=r, column=26, value=s2_ln)
        ws_sub.cell(row=r, column=27, value=s2_title)
        ws_sub.cell(row=r, column=28, value=s2_co)
        ws_sub.cell(row=r, column=29, value=s2_em)
        ws_sub.cell(row=r, column=30, value=s2_ft)
        ws_sub.cell(row=r, column=31, value=rating)
        ws_sub.cell(row=r, column=32, value=int_comm)
        ws_sub.cell(row=r, column=33, value=ext_comm)
        ws_sub.cell(row=r, column=34, value=sessiontype)
        ws_sub.cell(row=r, column=35, value=ev_code)
        ws_sub.cell(row=r, column=36, value=_submissions_print_formula(r))
        ws_sub.cell(row=r, column=37, value=_submissions_cpc_gridcell_formula(r))

    # Extra rows: special session slots (subset of columns; GridCell = IF title else tags)
    base_row = len(rows) + 2
    special_rows = _fetch_special_submission_rows(event_code)
    for j, spec in enumerate(special_rows):
        r = base_row + j
        slot, day, start_time, end_time, title, tags = spec
        ws_sub.cell(row=r, column=1, value=slot)
        ws_sub.cell(row=r, column=2, value=_submissions_special_gridcell_formula(r))
        ws_sub.cell(row=r, column=3, value=None)
        ws_sub.cell(row=r, column=4, value=title)
        ws_sub.cell(row=r, column=5, value=None)
        ws_sub.cell(row=r, column=6, value=tags)
        ws_sub.cell(row=r, column=7, value=None)
        ws_sub.cell(row=r, column=8, value=day)
        ws_sub.cell(row=r, column=9, value=_fmt_time_cell(start_time))
        ws_sub.cell(row=r, column=10, value=_fmt_time_cell(end_time))
        for c in range(11, 36):
            ws_sub.cell(row=r, column=c, value=None)
        ws_sub.cell(row=r, column=35, value=event_code)
        ws_sub.cell(row=r, column=37, value=f"=B{r}")
        ws_sub.cell(row=r, column=36, value=f"=AK{r}")

    n_data = len(rows) + len(special_rows)
    # Last row index for $A$2:$B$llll; if no data rows, use 2 so the range is valid
    return max(2, 1 + n_data)


def _write_print_sheet(ws_print, submissions_sheet_title, submissions_last_row):
    """
    Sheet print: column A references Submissions!AJ (Print column), same row count as submissions data.
    Column width ~4.88\"; row height ~1.80\"; left, indent 42, vertical center, wrap.
    """
    letter_a = get_column_letter(1)
    ws_print.column_dimensions[letter_a].width = PRINT_SHEET_COL_A_WIDTH_CHARS
    safe_title = str(submissions_sheet_title).replace("'", "''")
    qsheet = f"'{safe_title}'"
    if submissions_last_row < 2:
        return
    for r in range(2, submissions_last_row + 1):
        cell = ws_print.cell(row=r, column=1, value=f"={qsheet}!AJ{r}")
        cell.alignment = ALIGN_PRINT_SHEET_A
        ws_print.row_dimensions[r].height = PRINT_SHEET_ROW_HEIGHT_PT


@require_http_methods(["GET"])
def export_excel(request):
    """Export planner data to Excel. Returns .xlsx file as attachment."""
    today = date.today()
    event_code = _event_code(request)
    event_year_str = _event_year_for_export(request)
    version_num = _allocate_export_version(event_code, event_year_str)
    header_center = f"{event_code} {event_year_str} Grid"
    footer_center = f"Version {version_num}, {_export_footer_date_str(today)}"

    buffer = io.BytesIO()
    wb = Workbook()
    # Sheet 1: GridBySpeakerType (first/active sheet)
    ws_grid = wb.active
    ws_grid.title = "GridBySpeakerType"

    num_headers = _write_grid_column_headers(ws_grid, event_code)

    # Row 4: C4="Start", E4="End", then hide row 4; column F hidden
    ws_grid.cell(row=ROW_4_HEADER_ROW, column=3, value="Start")
    ws_grid.cell(row=ROW_4_HEADER_ROW, column=5, value="End")
    ws_grid.row_dimensions[ROW_4_HEADER_ROW].hidden = True
    ws_grid.column_dimensions[get_column_letter(COLUMN_F_INDEX)].hidden = True

    (
        last_calendar_row,
        last_calendar_col,
        date_rows,
        alternates_header_row,
    ) = _write_grid_row_headers(ws_grid, event_code, num_headers)

    # Sheet 2 first: grid formulas reference Submissions{yy}!$A$2:$B$llll
    year_suffix = str(date.today().year)[-2:]
    submissions_title = f"Submissions{year_suffix}"
    ws_sub = wb.create_sheet(title=submissions_title)
    submissions_last_row = _write_submissions_sheet(ws_sub, event_code)

    # Upgrade GridBySpeakerType: layout-driven cells (IFNA(VLOOKUP Submissions), merge colspan)
    _upgrade_grid_from_layout_sql(
        ws_grid, event_code, num_headers, submissions_title, submissions_last_row
    )
    _hide_grid_tags_helper_columns(ws_grid, num_headers)

    _apply_grid_alignment(ws_grid, num_headers)
    _apply_grid_track_submissions_tags_conditional_formatting(
        ws_grid, num_headers, submissions_title, submissions_last_row
    )
    _apply_grid_track_conditional_formatting(ws_grid, num_headers)
    # Track header G2:L3 when last_calendar_col missing (no row-header rows) use num_headers
    grid_last_col = (
        last_calendar_col
        if last_calendar_col is not None
        else COLUMN_HEADER_START_COL + num_headers - 1
    )
    # Borders: full outer frames first, then per-row B:E / G outlines (merge-aware)
    _apply_grid_track_header_outer_border(ws_grid, grid_last_col)
    _apply_grid_track_header_inner_verticals(ws_grid, grid_last_col)
    _apply_grid_calendar_outer_frame(ws_grid, last_calendar_row, last_calendar_col)
    _apply_grid_calendar_block_borders(
        ws_grid, last_calendar_row, last_calendar_col, date_rows
    )
    _apply_calendar_rows_extra_thin_outer_border(
        ws_grid, last_calendar_row, grid_last_col, date_rows
    )
    _apply_date_row_outline(ws_grid, date_rows, grid_last_col)
    ws_grid.page_setup.orientation = ws_grid.ORIENTATION_LANDSCAPE
    _apply_grid_sheet_header_footer(ws_grid, header_center, footer_center)
    _apply_grid_sheet_print_layout(
        ws_grid, grid_last_col, date_rows, alternates_header_row, event_code
    )

    # Sheet 3: print (column A = Submissions Print / AJ)
    ws_print = wb.create_sheet(title="print")
    _write_print_sheet(ws_print, submissions_title, submissions_last_row)

    # Sheet 4: CPC — VLOOKUP AK (CPCGridCell) + rating line; re-apply CF (copy_worksheet omits it)
    ws_cpc = wb.copy_worksheet(ws_grid)
    ws_cpc.title = "CPC"
    _apply_cpc_sheet_decorations(
        ws_cpc, event_code, num_headers, submissions_title, submissions_last_row
    )
    _apply_grid_alignment(ws_cpc, num_headers)
    _apply_grid_track_submissions_tags_conditional_formatting(
        ws_cpc, num_headers, submissions_title, submissions_last_row
    )
    _apply_grid_track_conditional_formatting(ws_cpc, num_headers)
    _apply_grid_sheet_header_footer(ws_cpc, header_center, footer_center)
    _apply_grid_sheet_print_layout(
        ws_cpc, grid_last_col, date_rows, alternates_header_row, event_code
    )

    ws_grid.sheet_view.zoomScale = EXCEL_SHEET_ZOOM_GRID_AND_SUBMISSIONS
    ws_sub.sheet_view.zoomScale = EXCEL_SHEET_ZOOM_GRID_AND_SUBMISSIONS
    ws_print.sheet_view.zoomScale = EXCEL_SHEET_ZOOM_PRINT
    ws_cpc.sheet_view.zoomScale = EXCEL_SHEET_ZOOM_GRID_AND_SUBMISSIONS

    wb.save(buffer)

    data = buffer.getvalue()
    filename = (
        f"IDUG_{event_code}{today.year}_Grid_{_export_filename_date_part(today)}_v{version_num}.xlsx"
    )

    response = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
